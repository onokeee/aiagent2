"""SELECT結果から Excel ブック(.xlsx)を組み立てる。

ファイルはディスクに書かず、メモリ上のバイト列として返す。

グラフはExcelネイティブのグラフとして入れる（画像ではない）。
受け取った側が範囲や種類を変えられるうえ、画像化ライブラリ（Chrome等）が
要らないので、サーバの環境に左右されない。
"""
from __future__ import annotations

import datetime as _dt  # noqa: F401  （シート値の型判定で使用）
import io
import re

from exports import XLSX_MIME  # noqa: F401  （既存の参照互換のため再公開）

from openpyxl import Workbook
from openpyxl.chart import (AreaChart, BarChart, LineChart, PieChart, Reference,
                            ScatterChart, Series)
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Excelのシート名に使えない文字と長さ制限
_BAD_SHEET_CHARS = re.compile(r"[\[\]:*?/\\]")
_SHEET_NAME_MAX = 31
_MAX_WIDTH = 60          # 列幅の上限（文字数）
_HEADER_FILL = PatternFill("solid", fgColor="1F3B5C")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_BAND_FILL = PatternFill("solid", fgColor="F5F8FC")
_THIN = Side(style="thin", color="D5DBE2")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# グラフの種類 -> (openpyxlのクラス, 積み上げ方)
CHART_TYPES = {
    "bar": (BarChart, "col", None),
    "bar_stacked": (BarChart, "col", "stacked"),
    "bar_percent": (BarChart, "col", "percentStacked"),
    "hbar": (BarChart, "bar", None),
    "hbar_stacked": (BarChart, "bar", "stacked"),
    "line": (LineChart, None, None),
    "line_stacked": (LineChart, None, "stacked"),
    "area": (AreaChart, None, None),
    "area_stacked": (AreaChart, None, "stacked"),
    "pie": (PieChart, None, None),
    "scatter": (ScatterChart, None, None),
}
_SERIES_COLORS = ["1F4E79", "F4B183", "70AD47", "C55A11", "7F7F7F",
                  "2E75B6", "A9D18E", "FFD966", "9DC3E6", "BFBFBF"]


def safe_sheet_name(name: str, used: set) -> str:
    """Excelの制約に合わせてシート名を整え、重複を避ける。"""
    s = _BAD_SHEET_CHARS.sub("_", str(name or "Sheet")).strip() or "Sheet"
    s = s[:_SHEET_NAME_MAX]
    base, i = s, 2
    while s.lower() in used:
        suffix = f"_{i}"
        s = base[: _SHEET_NAME_MAX - len(suffix)] + suffix
        i += 1
    used.add(s.lower())
    return s


def _cell_value(v):
    """openpyxl が扱えない型は文字列に落とす。"""
    if v is None or isinstance(v, (int, float, bool, str, _dt.datetime, _dt.date, _dt.time)):
        return v
    if isinstance(v, bytes):
        return f"<BLOB {len(v)} bytes>"
    return str(v)


def _autosize(ws, columns: list, rows: list):
    """見出しと先頭200行から列幅を決める。"""
    for ci, col in enumerate(columns, start=1):
        width = len(str(col))
        for r in rows[:200]:
            v = r[ci - 1] if ci - 1 < len(r) else None
            if v is not None:
                width = max(width, len(str(v)))
        ws.column_dimensions[get_column_letter(ci)].width = min(width + 2, _MAX_WIDTH)


def _add_chart(ws, spec: dict, columns: list, rows: list, header_row: int):
    """シートのデータ範囲からExcelネイティブのグラフを作って貼る。

    spec: {"type": 種類, "category_column": 横軸の列名, "value_columns": [系列の列名],
           "title": 見出し, "y_title": .., "x_title": .., "anchor": "H2",
           "data_labels": bool, "width": cm, "height": cm}
    """
    kind = str(spec.get("type") or "bar").lower()
    if kind not in CHART_TYPES:
        raise ValueError(f"未対応のグラフ種類です: {kind}。"
                         f"使えるのは {', '.join(CHART_TYPES)} です。")
    if not rows:
        raise ValueError("グラフにできる行がありません。")

    cat = spec.get("category_column") or (columns[0] if columns else None)
    if cat not in columns:
        raise ValueError(f"横軸の列 '{cat}' がありません。ある列: {', '.join(map(str, columns))}")
    vals = spec.get("value_columns") or [c for c in columns if c != cat]
    missing = [v for v in vals if v not in columns]
    if missing:
        raise ValueError(f"系列の列 {', '.join(map(str, missing))} がありません。"
                         f"ある列: {', '.join(map(str, columns))}")
    if not vals:
        raise ValueError("系列にする数値列がありません。")

    cls, direction, grouping = CHART_TYPES[kind]
    chart = cls()
    chart.title = spec.get("title") or None
    chart.style = 2
    if direction:
        chart.type = direction
    if grouping:
        chart.grouping = grouping
        chart.overlap = 100
    last = header_row + len(rows)
    cat_ref = Reference(ws, min_col=columns.index(cat) + 1, min_row=header_row + 1,
                        max_row=last)

    if kind == "scatter":
        # 散布図は x も数値列。1列目を x、残りを y にする。
        x_ref = Reference(ws, min_col=columns.index(vals[0]) + 1,
                          min_row=header_row + 1, max_row=last)
        for v in vals[1:] or vals[:1]:
            y_ref = Reference(ws, min_col=columns.index(v) + 1, min_row=header_row,
                              max_row=last)
            s = Series(y_ref, x_ref, title_from_data=True)
            s.marker.symbol = "circle"
            s.graphicalProperties.line.noFill = True
            chart.series.append(s)
    else:
        for i, v in enumerate(vals):
            ref = Reference(ws, min_col=columns.index(v) + 1, min_row=header_row,
                            max_row=last)
            chart.add_data(ref, titles_from_data=True)
        chart.set_categories(cat_ref)
        for i, s in enumerate(chart.series):
            color = _SERIES_COLORS[i % len(_SERIES_COLORS)]
            try:
                if kind.startswith("line"):
                    s.graphicalProperties.line.solidFill = color
                    s.smooth = False
                else:
                    s.graphicalProperties.solidFill = color
                    s.graphicalProperties.line.solidFill = color
            except AttributeError:
                pass

    if spec.get("data_labels") or (kind == "pie" and spec.get("data_labels") is not False):
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = kind != "pie"
        chart.dataLabels.showPercent = kind == "pie"
    if kind not in ("pie",):
        chart.y_axis.title = spec.get("y_title") or None
        chart.x_axis.title = spec.get("x_title") or None
        chart.y_axis.numFmt = spec.get("number_format") or "#,##0"
        chart.x_axis.delete = False      # これが無いとExcelで軸が消えることがある
        chart.y_axis.delete = False
    chart.width = float(spec.get("width") or 20)     # cm
    chart.height = float(spec.get("height") or 10)
    chart.legend.position = "b"
    if len(vals) <= 1 and kind != "pie":
        chart.legend = None

    anchor = spec.get("anchor") or f"{get_column_letter(len(columns) + 2)}{header_row}"
    ws.add_chart(chart, anchor)
    return chart


def build(sheets: list[dict], title: str | None = None) -> bytes:
    """[{"name", "columns", "rows", "note"?, "charts"?}, ...] から xlsx を作る。

    charts は同じシートのデータから作るグラフの指定（複数可）。
    """
    if not sheets:
        raise ValueError("シートが1つもありません。")
    wb = Workbook()
    wb.remove(wb.active)
    used: set = set()

    for sh in sheets:
        columns = list(sh.get("columns") or [])
        rows = list(sh.get("rows") or [])
        ws = wb.create_sheet(safe_sheet_name(sh.get("name"), used))

        start = 1
        note = str(sh.get("note") or "").strip()
        if note:
            ws.cell(row=1, column=1, value=note).font = Font(italic=True, color="666666")
            start = 3

        for ci, col in enumerate(columns, start=1):
            c = ws.cell(row=start, column=ci, value=str(col))
            c.font = _HEADER_FONT
            c.fill = _HEADER_FILL
            c.alignment = Alignment(vertical="center", horizontal="center")
            c.border = _BORDER
        for ri, row in enumerate(rows, start=start + 1):
            banded = (ri - start) % 2 == 0
            for ci in range(1, len(columns) + 1):
                cell = ws.cell(row=ri, column=ci,
                               value=_cell_value(row[ci - 1] if ci - 1 < len(row) else None))
                cell.border = _BORDER
                if banded:
                    cell.fill = _BAND_FILL
                if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                    cell.number_format = "#,##0.####"

        ws.freeze_panes = ws.cell(row=start + 1, column=1)
        if columns and rows:
            ws.auto_filter.ref = (f"A{start}:"
                                  f"{get_column_letter(len(columns))}{start + len(rows)}")
        _autosize(ws, columns, rows)

        charts = sh.get("charts")
        if isinstance(charts, dict):
            charts = [charts]
        for i, spec in enumerate(charts or []):
            spec = dict(spec or {})
            spec.setdefault("anchor",
                            f"{get_column_letter(len(columns) + 2)}"
                            f"{start + i * 21}")
            _add_chart(ws, spec, columns, rows, start)

    if title:
        wb.properties.title = str(title)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def safe_filename(name: str | None, default: str = "export") -> str:
    """互換用。実体は exports.safe_filename（拡張子 .xlsx）。"""
    from exports import safe_filename as _sf
    return _sf(name, "xlsx", default)
