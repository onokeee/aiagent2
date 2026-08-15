"""そのファイルが「そのまま取り込める表」かを判定する。

取り込みは 1行=1レコード / 1列=1項目 の素直な表を前提にしている。
ところが現場のExcelは、見出しがセル結合されていたり、月が横に並んでいたり、
合計行が混ざっていたりする。そのまま取り込むと、列名が「Unnamed: 3」になったり、
合計が二重に数えられたりして、後の集計が静かに狂う。

ここでは中身を読む前に形を見て、次のどれかを返す。
  そのまま取り込める / 手直しが要る / 取り込みに向かない / 対応していない形式

判定は「取り込みボタンを押す前に気づけるようにする」ためのもので、
最終的に決めるのは人。だから理由と直し方を必ず添える。
"""
from __future__ import annotations

import csv
import re
from pathlib import Path

import config

#: 形を見るために読む最大行数。これ以上は見なくても判断できる。
MAX_SCAN_ROWS = 200
#: 見出し行を探す範囲（先頭から何行目まで）。
HEADER_SEARCH_ROWS = 12
#: 結合セルの調査は通常読み込みが要る（メモリを食う）ので、この大きさまで。
MERGE_CHECK_MAX_MB = 20

VERDICTS = ("そのまま取り込める", "手直しが要る", "取り込みに向かない", "対応していない形式")

#: 合計・小計の行に出やすい言葉。混ざったまま取り込むと二重計上になる。
_TOTAL_WORDS = ("合計", "総計", "小計", "計", "累計", "total", "subtotal", "sum")
#: 見出しが日付・期間になっている＝横持ち（クロス表）の目印
_PERIOD_RE = re.compile(
    r"^\s*(?:"
    r"(?:19|20)\d{2}[-/年.]?(?:0?[1-9]|1[0-2])?[月]?"      # 2026-04 / 2026年4月
    r"|(?:0?[1-9]|1[0-2])月"                                # 4月
    r"|[QＱ][1-4]|第[1-4一二三四]四半期"                     # Q1 / 第1四半期
    r"|上期|下期|上半期|下半期"
    r")\s*$")


def _blank(v) -> bool:
    return v is None or str(v).strip() == ""


def _issue(level: str, text: str, fix: str = "") -> dict:
    return {"level": level, "text": text, "fix": fix}


# =============================================================================
# ファイルを「素の格子」として読む（見出しがどこかは、まだ決めつけない）
# =============================================================================

def _grid_excel(path: Path, sheet: str | None) -> tuple:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        names = list(wb.sheetnames)
        target = sheet if (sheet and sheet in names) else names[0]
        ws = wb[target]
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= MAX_SCAN_ROWS:
                break
            rows.append(list(row))
        return rows, names, target
    finally:
        wb.close()


def _merged_ranges(path: Path, sheet: str) -> list | None:
    """結合セルの範囲。読み取り専用モードでは取れないので通常読み込みする。

    大きいファイルで開くと重いので、その場合は調べずに None を返す
    （「分からなかった」と「無かった」を混同しないため）。
    """
    try:
        if path.stat().st_size > MERGE_CHECK_MAX_MB * 1024 * 1024:
            return None
    except OSError:
        return None
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
        try:
            ws = wb[sheet] if sheet in wb.sheetnames else wb[wb.sheetnames[0]]
            return [(r.min_row, r.min_col, r.max_row, r.max_col)
                    for r in ws.merged_cells.ranges]
        finally:
            wb.close()
    except Exception:
        return None


def _grid_text(path: Path) -> tuple:
    """CSV/TSV/TXT。区切り文字と文字コードもここで見当をつける。"""
    from importer import CSV_ENCODINGS

    raw = None
    for enc in CSV_ENCODINGS:
        try:
            raw = path.read_text(encoding=enc)
            used = enc
            break
        except (UnicodeDecodeError, OSError):
            continue
    if raw is None:
        raise ValueError("文字コードを判定できませんでした（UTF-8 か Shift_JIS で保存し直してください）。")

    head = "\n".join(raw.splitlines()[:MAX_SCAN_ROWS])
    if path.suffix.lower() == ".tsv":
        delim = "\t"
    else:
        try:
            delim = csv.Sniffer().sniff(head[:4000], delimiters=",\t;|").delimiter
        except csv.Error:
            delim = "\t" if head.count("\t") > head.count(",") else ","
    rows = [r for r in csv.reader(head.splitlines(), delimiter=delim)]
    return rows, used, delim


# =============================================================================
# 形を見る
# =============================================================================

def _guess_header(rows: list) -> int:
    """見出しの行番号（0始まり）を当てる。

    「文字が並んでいて、その下に中身が続いている行」を見出しとみなす。
    タイトル行（1セルだけ埋まっている）や空行は飛ばす。
    """
    best, best_score = 0, -1.0
    for i, row in enumerate(rows[:HEADER_SEARCH_ROWS]):
        filled = [v for v in row if not _blank(v)]
        if len(filled) < 2:
            continue                       # タイトル行や空行
        below = rows[i + 1] if i + 1 < len(rows) else []
        if not [v for v in below if not _blank(v)]:
            continue                       # 下に中身が無いなら見出しではない
        texty = sum(1 for v in filled if not _looks_number(v))
        # 埋まり具合＋文字らしさ。上の行ほど見出しらしいので少し優遇する
        score = (len(filled) / max(len(row), 1)) + (texty / len(filled)) - i * 0.06
        if score > best_score:
            best, best_score = i, score
    return best


def _looks_number(v) -> bool:
    s = str(v).strip().replace(",", "")
    if not s:
        return False
    try:
        float(s)
        return True
    except ValueError:
        return False


def _is_total_row(row: list) -> bool:
    head = " ".join(str(v) for v in row[:2] if not _blank(v)).strip().lower()
    return bool(head) and any(w in head for w in _TOTAL_WORDS)


def _analyze(rows: list, header_row: int) -> dict:
    """見出し行を決めた上で、中身の形を調べる。"""
    header = rows[header_row] if header_row < len(rows) else []
    body = rows[header_row + 1:]
    width = max((len(r) for r in rows), default=0)

    names = [("" if _blank(v) else str(v).strip()) for v in header]
    names += [""] * (width - len(names))

    empty_names = [i for i, n in enumerate(names) if not n]
    dup = sorted({n for n in names if n and names.count(n) > 1})
    period_cols = [n for n in names if n and _PERIOD_RE.match(n)]
    numeric_names = [n for n in names if n and _looks_number(n)]

    blank_rows = sum(1 for r in body if all(_blank(v) for v in r))
    total_rows = [i for i, r in enumerate(body) if _is_total_row(r)]
    ragged = sum(1 for r in body if len([v for v in r if not _blank(v)]) > len(names))
    multiline = any(isinstance(v, str) and "\n" in v for r in body[:50] for v in r)

    # 全部空の列（見出しだけあって中身が無い／見出しも中身も無い）
    empty_cols = []
    for c in range(width):
        col = [r[c] for r in body if c < len(r)]
        if col and all(_blank(v) for v in col):
            empty_cols.append(names[c] or f"{c + 1}列目")

    # 数字と文字が混ざる列（"-" や "N/A" が入ると、数値として取り込めない）
    mixed = []
    for c in range(width):
        col = [r[c] for r in body if c < len(r) and not _blank(r[c])]
        if len(col) < 4:
            continue
        nums = sum(1 for v in col if _looks_number(v))
        if 0.6 <= nums / len(col) < 1.0:
            odd = [str(v) for v in col if not _looks_number(v)][:3]
            mixed.append((names[c] or f"{c + 1}列目", odd))

    return {
        "names": names, "width": width, "body_rows": len(body),
        "empty_names": empty_names, "dup_names": dup,
        "period_cols": period_cols, "numeric_names": numeric_names,
        "blank_rows": blank_rows, "total_rows": total_rows,
        "ragged": ragged, "multiline": multiline,
        "empty_cols": empty_cols, "mixed": mixed,
    }


def _blocks(rows: list, header_row: int) -> int:
    """1シートに表がいくつ入っていそうか（空行で切れて、また見出しが始まる）。"""
    blocks, in_block, gap = 1, True, 0
    for r in rows[header_row + 1:]:
        if all(_blank(v) for v in r):
            gap += 1
            in_block = False
        else:
            if not in_block and gap >= 2:
                blocks += 1
            in_block, gap = True, 0
    return blocks


# =============================================================================
# 判定
# =============================================================================

def inspect(path, sheet: str | None = None) -> dict:
    """1ファイル（Excelは1シート）の形を見て、取り込めるかを判定する。"""
    p = Path(path)
    ext = p.suffix.lower()
    out = {"file": p.name, "sheet": sheet, "sheets": [], "header_row": 0,
           "verdict": "", "issues": [], "shape": {}}

    if ext not in config.IMPORT_EXTENSIONS:
        out["verdict"] = "対応していない形式"
        out["issues"] = [_issue(
            "高", f"{ext or '拡張子なし'} は取り込みに対応していません。",
            f"扱えるのは {'、'.join(config.IMPORT_EXTENSIONS)} です。"
            "元のシステムからCSVで出し直すか、Excelで開いて「名前を付けて保存」で"
            ".xlsx か .csv にしてください。")]
        return out

    try:
        if ext in (".xlsx", ".xlsm"):
            rows, names, target = _grid_excel(p, sheet)
            out["sheets"], out["sheet"] = names, target
            merged = _merged_ranges(p, target)
        else:
            rows, enc, delim = _grid_text(p)
            merged = []
            out["encoding"] = enc
            out["delimiter"] = {"\t": "タブ", ",": "カンマ", ";": "セミコロン",
                                "|": "パイプ"}.get(delim, delim)
    except Exception as e:
        out["verdict"] = "取り込みに向かない"
        out["issues"] = [_issue("高", f"ファイルを開けませんでした: {e}",
                                "壊れているか、パスワードが掛かっている可能性があります。")]
        return out

    rows = [r for r in rows if r is not None]
    if not any(any(not _blank(v) for v in r) for r in rows):
        out["verdict"] = "取り込みに向かない"
        out["issues"] = [_issue("高", "中身が空です。", "データの入ったファイルを指定してください。")]
        return out

    header_row = _guess_header(rows)
    info = _analyze(rows, header_row)
    blocks = _blocks(rows, header_row)
    out["header_row"] = header_row
    out["shape"] = {"列数": info["width"], "読んだ行数": info["body_rows"],
                    "見出し行": header_row + 1}
    out["columns"] = info["names"]

    issues: list[dict] = []

    # --- そのままでは取り込めないもの ---------------------------------------
    if merged:
        hrow = header_row + 1                    # 1始まりの行番号にそろえる
        # 見出しの行と、そのすぐ上をまたぐ横方向の結合＝多段見出し。
        # いちばん上のタイトル行（1セルだけの飾り）は、これに含めない。
        in_header = [m for m in merged
                     if m[1] != m[3] and m[0] <= hrow and m[2] >= hrow - 1
                     and not (m[0] == m[2] == 1 and hrow > 2)]
        in_body = [m for m in merged if m[0] > hrow]
        if in_header:
            issues.append(_issue(
                "高", f"見出しがセル結合されています（{len(in_header)}箇所）。"
                      "多段の見出しは1行の列名にできません。",
                "結合を解除し、見出しを1行にまとめてください"
                "（例:「上期／4月」→「上期_4月」）。"))
        if in_body:
            issues.append(_issue(
                "高", f"データ部分にセル結合があります（{len(in_body)}箇所）。"
                      "結合されたセルは先頭以外が空になり、行が正しく揃いません。",
                "結合を解除し、空いたセルに同じ値を埋めてください。"))

    if len(info["period_cols"]) >= 3:
        issues.append(_issue(
            "高", f"月や期間が横に並んでいます（{'、'.join(info['period_cols'][:5])}…）。"
                  "いわゆるクロス表で、1行=1レコードになっていません。",
            "「年月」「値」の2列に縦持ちへ直してください"
            "（Excelなら [データ]→[パワークエリ]→[列のピボット解除]）。"))
    elif len(info["numeric_names"]) >= 3:
        issues.append(_issue(
            "高", f"見出しが数字になっています（{'、'.join(info['numeric_names'][:5])}…）。"
                  "見出し行の位置が違うか、横持ちの表の可能性があります。",
            "1行目に列名が来るようにしてください。"))

    if blocks > 1:
        issues.append(_issue(
            "高", f"1つのシートに表が{blocks}個あるように見えます（間に空行があります）。",
            "表ごとにシートを分けてください。取り込みは1シート=1テーブルです。"))

    # --- 直せば取り込めるもの ------------------------------------------------
    if header_row > 0:
        issues.append(_issue(
            "中", f"{header_row + 1}行目が見出しに見えます（1行目ではありません）。"
                  "上にタイトルや空行が入っています。",
            f"取り込み画面の「見出しの行」に {header_row + 1} を指定するか、"
            "上の行を削除してください。"))

    if info["total_rows"]:
        issues.append(_issue(
            "中", f"合計・小計らしい行が {len(info['total_rows'])} 行あります。"
                  "そのまま取り込むと二重に数えられます。",
            "合計行を削除してから取り込んでください（集計はアプリ側でできます）。"))

    if info["empty_names"]:
        issues.append(_issue(
            "中", f"列名が空の列が {len(info['empty_names'])} 個あります。",
            "列名を付けてください（空のままだと自動で仮の名前が付きます）。"))

    if info["dup_names"]:
        issues.append(_issue(
            "中", f"同じ列名が複数あります: {'、'.join(info['dup_names'][:5])}",
            "区別できる名前に変えてください（取り込み時は連番が付きます）。"))

    if info["ragged"]:
        issues.append(_issue(
            "中", f"見出しより列が多い行が {info['ragged']} 行あります。"
                  "区切り文字がデータの中に入っている可能性があります。",
            "その列を引用符で囲むか、区切り文字を変えて出し直してください。"))

    # --- 気に留めておく程度 --------------------------------------------------
    for name, odd in info["mixed"][:3]:
        issues.append(_issue(
            "低", f"「{name}」は数字の列に見えますが、文字が混ざっています"
                  f"（{'、'.join(odd)}）。",
            "空欄や「-」「N/A」は空にしておくと、数値として取り込めます。"))
    if info["empty_cols"]:
        issues.append(_issue(
            "低", f"中身が空の列があります: {'、'.join(info['empty_cols'][:5])}",
            "取り込む列の選択から外せます。"))
    if info["blank_rows"]:
        issues.append(_issue(
            "低", f"途中に空行が {info['blank_rows']} 行あります。", "空行は取り込み時に残ります。"))
    if info["multiline"]:
        issues.append(_issue(
            "低", "セルの中で改行しているところがあります。",
            "表示は崩れませんが、検索や集計がしにくくなります。"))
    if merged is None:
        issues.append(_issue(
            "低", "ファイルが大きいため、セル結合までは調べていません。", ""))

    levels = {i["level"] for i in issues}
    out["issues"] = issues
    out["verdict"] = ("取り込みに向かない" if "高" in levels else
                      "手直しが要る" if "中" in levels else
                      "そのまま取り込める")
    return out


def summary_line(res: dict) -> str:
    """一覧に出す一言。"""
    high = [i for i in res["issues"] if i["level"] == "高"]
    if res["verdict"] == "そのまま取り込める":
        return "そのまま取り込める"
    if high:
        return f"{res['verdict']}（{high[0]['text'][:40]}）"
    mid = [i for i in res["issues"] if i["level"] == "中"]
    return f"{res['verdict']}（{mid[0]['text'][:40]}）" if mid else res["verdict"]
